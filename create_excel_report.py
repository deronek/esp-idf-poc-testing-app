import glob
import os
import time
from datetime import datetime

from junitparser import TestCase, JUnitXml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils import is_jenkins

# Constants
TEST_REPORT_SHEET_NAME = "Test report"
REPORT_HEADER_START_ROW = 2
REPORT_HEADER_END_ROW = 8

RESULTS_COUNT_START_ROW = 11
RESULTS_COUNT_END_ROW = 14

DATA_HEADER_ROW = 17
TEST_DATA_START_ROW = 18

TEST_CASE_COLUMN_INDEX = 1
STATUS_COLUMN_INDEX = 2
DURATION_COLUMN_INDEX = 3

DEFECTS_SHEET_NAME = "Found defects"
DEFECTS_TABLE_HEADER_ROW = 1
DEFECTS_TABLE_START_ROW = 2
DEFECTS_TEST_CASE_COLUMN_INDEX = 1
DEFECTS_DESCRIPTION_COLUMN_INDEX = 2

COLOR_MAPPINGS = {
    'Passed': "00FF00",  # Green
    'Failed': "FF0000",  # Red
    'Skipped': "D3D3D3"  # Grey
}


def create_excel_report(xml_files, output_file):
    wb = Workbook()
    ws_test_report = wb.active
    ws_test_report.title = TEST_REPORT_SHEET_NAME

    # Add header information to the "Test report" sheet
    add_header_info(ws_test_report)
    fill_jenkins_header_data(ws_test_report)

    ws_test_report.append(['Test Case', 'Status', 'Duration'])

    for xml_file in xml_files:
        junit_xml = JUnitXml.fromfile(xml_file)
        for suite in junit_xml:
            for case in suite:
                append_test_case_data(ws_test_report, case)

    apply_styles_to_worksheet(ws_test_report)
    apply_styles_to_header(ws_test_report)
    adjust_column_widths(ws_test_report)  # Adjust column widths after content is added
    center_text_in_cells(ws_test_report)  # Center text horizontally and vertically
    format_duration_column(ws_test_report)  # Format "Duration" column as seconds with 2 decimal places

    # Update formulas for counting Passed/Failed/Skipped test cases
    update_formulas_for_test_counts(ws_test_report)

    # Create "Found defects" sheet
    ws_defects = wb.create_sheet(title=DEFECTS_SHEET_NAME)
    create_defects_table(ws_defects)
    adjust_column_widths(ws_defects)

    wb.save(output_file)


def update_formulas_for_test_counts(ws):
    last_row = ws.max_row
    ws['B11'] = f'=COUNTIF(B{TEST_DATA_START_ROW}:B{last_row}, "<>")'
    ws['B12'] = f'=COUNTIF(B{TEST_DATA_START_ROW}:B{last_row}, "Passed")'
    ws['B13'] = f'=COUNTIF(B{TEST_DATA_START_ROW}:B{last_row}, "Failed")'
    ws['B14'] = f'=COUNTIF(B{TEST_DATA_START_ROW}:B{last_row}, "Skipped")'


def fill_jenkins_header_data(ws):
    if is_jenkins():
        ws['B2'] = os.getenv('JOB_NAME').split('/')[0]
        ws['B3'] = 'Proof-of-concept app for Jenkins HIL testing workflow'
        ws['B4'] = os.getenv('BUILD_USER')
        ws['B5'] = os.getenv('BUILD_USER_EMAIL')
        ws['B7'] = os.getenv('SW_VARIANT')
        ws['B8'] = os.getenv('HW_VARIANT')


def add_header_info(ws: Worksheet):
    name_style = NamedStyle(name="header_name_style", font=Font(bold=True, size=14),
                            alignment=Alignment(horizontal="center"))
    value_style = NamedStyle(name="header_value_style", font=Font(size=14), alignment=Alignment(horizontal="center"))

    header_info = [
        ["Application Name:", ""],
        ["Project Name:", ""],
        ["Engineer Name:", ""],
        ["Engineer E-mail:", ""],
        ["Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["SW Variant:", ""],
        ["HW Variant:", ""]
    ]
    for name, value in header_info:
        name_cell = ws.cell(row=ws.max_row + 1, column=1, value=name)
        value_cell = ws.cell(row=ws.max_row, column=2, value=value)
        name_cell.style = name_style
        value_cell.style = value_style

    # Header of result count - Passed/Failed/Skipped test cases
    ws.cell(row=ws.max_row + 3, column=1, value="All")
    ws.cell(row=ws.max_row + 1, column=1, value="Passed")
    ws.cell(row=ws.max_row + 1, column=1, value="Failed")
    ws.cell(row=ws.max_row + 1, column=1, value="Skipped")
    ws.append([])
    ws.append([])


def append_test_case_data(ws, case):
    case_name = case.name if case.name else "N/A"
    case_result = get_str_result(case)
    case_time = case.time if case.time else 0.0

    ws.append([case_name, case_result, case_time])


def apply_styles_to_worksheet(ws):
    apply_color_mapping_styles(ws, RESULTS_COUNT_START_ROW, RESULTS_COUNT_END_ROW, 1)
    apply_color_mapping_styles(ws, TEST_DATA_START_ROW, ws.max_row, STATUS_COLUMN_INDEX)

    # Apply bold font to Passed/Failed/Skipped headers in results count
    for row in ws.iter_rows(min_row=RESULTS_COUNT_START_ROW, max_row=RESULTS_COUNT_END_ROW, min_col=1,
                            max_col=1):
        for cell in row:
            cell.font = Font(size=12, bold=True)


def apply_color_mapping_styles(ws, start_row, end_row, column):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=column,
                            max_col=column):
        case_result_cell = row[0]
        case_result = case_result_cell.value
        if case_result in COLOR_MAPPINGS:
            cell_fill = PatternFill(start_color=COLOR_MAPPINGS[case_result], end_color=COLOR_MAPPINGS[case_result],
                                    fill_type="solid")
            case_result_cell.fill = cell_fill


def apply_styles_to_header(ws):
    header_row = ws[DATA_HEADER_ROW]
    header_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # Dark Grey
    header_font = Font(size=12, bold=True, color="FFFFFF")  # White

    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font


def adjust_column_widths(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Get the column letter of the first cell
        for cell in column:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add a buffer and multiply by a factor for width adjustment
        ws.column_dimensions[column_letter].width = adjusted_width


def center_text_in_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def format_duration_column(ws):
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row,
                            min_col=DURATION_COLUMN_INDEX, max_col=DURATION_COLUMN_INDEX):
        duration_cell = row[0]
        duration_value = duration_cell.value
        if isinstance(duration_value, (int, float)):
            duration_cell.number_format = '0.00" s"'


def create_defects_table(ws):
    header_style = NamedStyle(name="defects_table_header_style", font=Font(bold=True),
                              alignment=Alignment(horizontal="center"))
    table_header = ["Failed test case", "Description"]

    ws.append(table_header)
    ws.freeze_panes = ws.cell(row=2, column=1)
    for cell in ws[DEFECTS_TABLE_HEADER_ROW]:
        cell.style = header_style


def get_str_result(case: TestCase) -> str:
    if case.is_passed:
        return 'Passed'
    elif case.is_skipped:
        return 'Skipped'
    else:
        return 'Failed'


if __name__ == "__main__":
    workspace_path = os.getenv("WORKSPACE")
    xml_files = glob.glob(os.getenv("WORKSPACE") + r"\*.xml")
    output_file = os.path.join(workspace_path, f'Test_Report_{time.strftime("%Y-%m-%d_%Hh%Mm%Ss")}.xlsx')
    create_excel_report(xml_files, output_file)
